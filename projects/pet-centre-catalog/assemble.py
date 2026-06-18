#!/usr/bin/env python
"""Write enriched descriptions + image URLs back into the catalog.

Two modes:
  --preview : build a compact standalone workbook from a subset (for review).
  (default) : copy the source workbook and fill columns B (Description), F (Image URL),
              and a new G (Image Confidence), preserving every other column.

Image cells are colour-coded so curation is fast and a wrong guess can never look
confirmed: VERIFIED = clean (white), LIKELY = red fill ("check me"), BLANK = grey.

Usage:
  python assemble.py --rows rows.json --desc desc.json --img img.json --out out.xlsx [--src src.xlsx] [--preview]
"""
import argparse
import io
import json
import os
import urllib.request
from urllib.parse import urlsplit

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None


def fetch_thumb(url, box=110):
    """Download an image and return a small PNG thumbnail in a BytesIO (preview only)."""
    if not (url and PILImage):
        return None
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=20) as r:
            data = r.read()
        im = PILImage.open(io.BytesIO(data)).convert("RGB")
        im.thumbnail((box, box))
        out = io.BytesIO()
        im.save(out, format="PNG")
        out.seek(0)
        return out
    except Exception:
        return None


def thumb_for(row, url, imgdir, box=110):
    """Prefer a pre-normalized local image (<imgdir>/<row>.jpg) for the embedded thumbnail; else the URL."""
    if imgdir and PILImage:
        p = os.path.join(imgdir, f"{row}.jpg")
        if os.path.exists(p):
            try:
                im = PILImage.open(p).convert("RGB")
                im.thumbnail((box, box))
                out = io.BytesIO()
                im.save(out, format="PNG")
                out.seek(0)
                return out
            except Exception:
                pass
    return fetch_thumb(url)

YELLOW = PatternFill("solid", fgColor="FFEB9C")   # likely / unverified — review this
RED = PatternFill("solid", fgColor="FFC7CE")      # blank — needs manual sourcing
GREEN = PatternFill("solid", fgColor="C6EFCE")    # verified (any tier) — whole row
HEAD = PatternFill("solid", fgColor="00975A")     # Pet Centre green
HEADFONT = Font(bold=True, color="FFFFFF")
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")  # wrap long text in-cell, left-align
LEFT_TOP = Alignment(vertical="top", horizontal="left")  # left-align numbers too (not Excel's default right)


def load(path):
    return json.load(open(path, encoding="utf-8"))


def img_url(rec):
    return rec.get("url", "") if rec else ""


# Human-readable col-G labels: every verified tier says HOW it was confirmed.
VERIFIED_LABELS = {
    "verified-official": "verified (official site)",  # barcode confirmed on the brand's own site
    "verified-cross": "verified (2+ sources)",        # barcode confirmed on 2+ independent sites
    "verified": "verified (barcode)",                 # barcode confirmed on a trusted page
    "verified-visual": "verified (image)",            # AI confirmed the photo is the product
}


def confidence(rec):
    c = rec.get("confidence") if rec else None
    if c in VERIFIED_LABELS:
        return VERIFIED_LABELS[c]
    if c == "likely":
        return "likely (check)"
    return "blank: " + ((rec or {}).get("reason") or "no-data")


def _img_src(rec):
    """Where the PICTURE came from, in plain words."""
    if not rec or not rec.get("url"):  # rec is None for rows not yet resolved (capped/partial run)
        return "none"
    q = rec.get("img_quality") or ""
    if q:  # couldn't get a clean photo — flag for manual swap
        return "check - " + {"multi": "multi-product", "crop": "cropped", "unclear": "unclear"}.get(q, q)
    prov = rec.get("img_provenance") or ""
    if prov.startswith("off-source:"):
        return "off-source (" + prov.split(":", 1)[1] + ")"
    if prov == "ai-matched":
        return "AI-matched"
    if not (rec.get("confidence") or "").startswith("verified"):
        return "best-guess"  # likely tier — image is unconfirmed
    return "brand site" if prov == "official" else "verified source"


def _txt_src(rec):
    """Where the DESCRIPTION facts came from, in plain words."""
    if not rec:  # rec is None for rows not yet resolved (capped/partial run)
        return "name-only"
    dp = rec.get("desc_provenance") or ""
    if dp.startswith("supplemented:"):
        return "+" + dp.split(":", 1)[1]
    if dp == "source":
        return "verified source"
    return "name-only"  # no page text (likely/blank) — generic from the name/category


def source_notes(rec):
    """Spell out BOTH provenances every row so it's unambiguous which is which:
    IMG: where the picture came from   ·   TXT: where the description facts came from."""
    if not rec:
        return "IMG: none  -  TXT: name-only"
    return f"IMG: {_img_src(rec)}  -  TXT: {_txt_src(rec)}"


# --- Hike-upload sheet structure (one sheet; Hike maps columns by HEADER NAME) -----------------
UPLOAD_HEADERS = ["Name", "Description", "Barcode", "Product Type", "Brand Name", "Image URL",
                  "Depth", "Width", "Height", "Weight"]   # named EXACTLY as Hike's import fields
DIVIDER_HEADER = "‖ DO NOT IMPORT →"             # visual fence; left unmapped on import
QA_HEADERS = ["Status", "Image Confidence", "Image Source", "Text Source", "Image Flag", "Best guess (manual)"]
ALL_HEADERS = UPLOAD_HEADERS + [DIVIDER_HEADER] + QA_HEADERS
IMG_COL = 6                                # Image URL column (1-based) — carries the confidence colour
DIVIDER_COL = len(UPLOAD_HEADERS) + 1      # column K


def desc_obj(desc, r):
    """desc[r] may be a structured object {description, depth, ...} or a legacy plain string."""
    v = desc.get(r)
    if isinstance(v, dict):
        return v
    if isinstance(v, str):
        return {"description": v}
    return {}


def ready(rec):
    """Owner rule: push Verified + Likely; hold blank/no-match for manual sourcing."""
    c = (rec or {}).get("confidence") or ""
    return "READY" if (c.startswith("verified") or c == "likely") else "HOLD"


def img_flag(rec):
    q = (rec or {}).get("img_quality") or ""
    return {"multi": "multi-product", "crop": "cropped", "unclear": "unclear"}.get(q, "")


def dims_cells(dobj):
    """Depth/Width/Height/Weight cell values (blank when not stated)."""
    return [dobj.get("depth") or "", dobj.get("width") or "",
            dobj.get("height") or "", dobj.get("weight") or ""]


def qa_cells(rec):
    """The fenced curation columns, in order (the last is linked out in set_best_guess)."""
    return [ready(rec), confidence(rec), _img_src(rec), _txt_src(rec), img_flag(rec), best_guess_text(rec)]


TXT_SRC_COL = ALL_HEADERS.index("Text Source") + 1   # 1-based column made into a clickable link


def txt_src_url(rec):
    """The source PAGE url the description facts came from — for a clickable Text Source (mirrors the
    image link). Empty for name-only rows (no page text scraped, so nothing to link to)."""
    if not rec:
        return ""
    dp = rec.get("desc_provenance") or ""
    if dp == "source" or dp.startswith("supplemented:"):
        return rec.get("src") or ""
    return ""


def row_fill(rec):
    """Whole-row tier colour: verified -> green, likely -> yellow, blank/no-match -> red."""
    c = (rec or {}).get("confidence") or ""
    if c.startswith("verified"):
        return GREEN
    if c == "likely":
        return YELLOW
    return RED


def link_cell(cell, url):
    """Render a cell as a clickable blue hyperlink when url is non-empty (else leave as-is)."""
    if url:
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")


BEST_COL = ALL_HEADERS.index("Best guess (manual)") + 1


def best_guess_text(rec):
    """For a blank/unconfirmed row, the closest candidate title we found — a manual-sourcing head
    start. Empty for verified/likely rows (they already carry a confident result)."""
    return "" if (not rec or rec.get("confidence")) else (rec.get("best_title") or "")


def best_guess_url(rec):
    return "" if (not rec or rec.get("confidence")) else (rec.get("best_url") or "")


def src_domain(url):
    """Bare host for at-a-glance source identity, e.g. 'shop.megapet.com.au' (drops a leading www.)."""
    host = urlsplit(url).netloc.lower()
    return host[4:] if host.startswith("www.") else host


def set_text_source(ws, r, rec):
    """Text Source shows the source's NAME (domain) as a clickable link so provenance reads at a
    glance; name-only rows keep their plain label (from qa_cells), no link."""
    turl = txt_src_url(rec)
    if turl:
        link_cell(ws.cell(row=r, column=TXT_SRC_COL, value=src_domain(turl)), turl)


def set_best_guess(ws, r, rec):
    """Make the blank-row best-guess title clickable (to the closest candidate we found)."""
    link_cell(ws.cell(row=r, column=BEST_COL), best_guess_url(rec))


# Per-column display: width auto-fits to the widest value (capped here); wrapped columns spill onto
# extra lines in-cell instead of bleeding into the next column. Indexed 1-based to the sheet columns.
COL_DISPLAY = {
    1: (42, True),    # Name
    2: (58, True),    # Description
    3: (16, False),   # Barcode
    4: (26, True),    # Product Type
    5: (20, True),    # Brand Name
    6: (46, True),    # Image URL
    7: (9, False), 8: (9, False), 9: (9, False), 10: (10, False),  # Depth/Width/Height/Weight
    11: (18, False),  # divider
    12: (9, False),   # Status
    13: (22, True),   # Image Confidence
    14: (24, True),   # Image Source
    15: (26, True),   # Text Source
    16: (15, False),  # Image Flag
    17: (40, True),   # Best guess (manual)
}


def fit_and_wrap(ws, ncols, last_row):
    """Auto-fit each column to its widest value (capped per COL_DISPLAY) and wrap the long-text
    columns, so no cell spills into the next column. Excel auto-grows row heights for wrapped text."""
    for c in range(1, ncols + 1):
        cap, wrap = COL_DISPLAY.get(c, (24, False))
        longest = 0
        for r in range(1, last_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and len(str(v)) > longest:
                longest = len(str(v))
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = min(max(longest + 2, 8), cap)
        align = WRAP_TOP if wrap else LEFT_TOP  # numbers left-aligned, not Excel's default right
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=c).alignment = align


# Curation (review) sheet column order: the image is column A, then everything you verify grouped
# left-to-right (identity + description + dims first, then the confidence/provenance aids). This is the
# HUMAN working copy; the clean Hike-import sheet is a separate extraction built later on request.
# Each entry: (header, value_fn(p, rec, dobj), link_fn(rec)->url|None, width_cap, wrap?).
CURATION_COLS = [
    ("Name",             lambda p, rec, d: p["name"],                None,                      34, True),
    ("Brand Name",       lambda p, rec, d: p["brand"],               None,                      18, True),
    ("Product Type",     lambda p, rec, d: p["type"],                None,                      22, True),
    ("Barcode",          lambda p, rec, d: p["barcode"],             None,                      16, False),
    ("Description",      lambda p, rec, d: d.get("description", ""),  None,                      58, True),
    ("Depth (cm)",       lambda p, rec, d: d.get("depth") or "",      None,                      11, False),
    ("Width (cm)",       lambda p, rec, d: d.get("width") or "",       None,                      11, False),
    ("Height (cm)",      lambda p, rec, d: d.get("height") or "",      None,                      11, False),
    ("Weight (kg)",      lambda p, rec, d: d.get("weight") or "",      None,                      12, False),
    ("Status",           lambda p, rec, d: ready(rec),               None,                      9,  False),
    ("Image Confidence", lambda p, rec, d: confidence(rec),          None,                      22, True),
    ("Image Flag",       lambda p, rec, d: img_flag(rec),            None,                      15, False),
    ("Image URL",        lambda p, rec, d: img_url(rec),             lambda rec: img_url(rec),   44, True),
    ("Image Source",     lambda p, rec, d: _img_src(rec),            None,                      22, True),
    ("Text Source",      lambda p, rec, d: src_domain(txt_src_url(rec)) if txt_src_url(rec) else _txt_src(rec),
                                                                     lambda rec: txt_src_url(rec), 24, True),
    ("Best guess (manual)", lambda p, rec, d: best_guess_text(rec),  lambda rec: best_guess_url(rec), 40, True),
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", required=True)
    ap.add_argument("--desc", required=True)
    ap.add_argument("--img", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--src")
    ap.add_argument("--preview", action="store_true")
    ap.add_argument("--embed", action="store_true", help="embed image thumbnails (preview only)")
    ap.add_argument("--imgdir", help="use pre-normalized local images (<row>.jpg) for embedded thumbnails")
    a = ap.parse_args()

    rows = load(a.rows)
    desc = {int(k): v for k, v in load(a.desc).items()}
    img = {int(k): v for k, v in load(a.img).items()}

    if a.preview:
        # Curation workbook: image in column A, verification fields grouped left-to-right (CURATION_COLS).
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Curation"
        headers = ["Image"] + [c[0] for c in CURATION_COLS]
        ws.append(headers)
        for i in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=i)
            cell.fill = HEAD
            cell.font = HEADFONT
        for p in rows:
            rec = img.get(p["row"])
            dobj = desc_obj(desc, p["row"])
            ws.append([""] + [vfn(p, rec, dobj) for _, vfn, _, _, _ in CURATION_COLS])
            r = ws.max_row
            for j, (_, _, lfn, _, _) in enumerate(CURATION_COLS, start=2):
                if lfn:
                    link_cell(ws.cell(row=r, column=j), lfn(rec))
            fill = row_fill(rec)                             # whole-row tier colour
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = fill
            if a.embed:  # embed the product thumbnail in column A for at-a-glance confirmation
                thumb = thumb_for(p["row"], img_url(rec), a.imgdir)
                if thumb:
                    ws.add_image(XLImage(thumb), f"A{r}")
                    ws.row_dimensions[r].height = 85
        ws.column_dimensions["A"].width = 16  # image column
        for j, (_, _, _, cap, wrap) in enumerate(CURATION_COLS, start=2):  # auto-fit to content, capped
            longest = len(str(headers[j - 1]))
            for rr in range(2, ws.max_row + 1):
                v = ws.cell(row=rr, column=j).value
                if v is not None and len(str(v)) > longest:
                    longest = len(str(v))
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = min(max(longest + 2, 8), cap)
            align = WRAP_TOP if wrap else LEFT_TOP  # numbers left-aligned, not Excel's default right
            for rr in range(2, ws.max_row + 1):
                ws.cell(row=rr, column=j).alignment = align
        ws.freeze_panes = "B2"   # keep the image column + header visible while scrolling
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{ws.max_row}"
        wb.save(a.out)
        print(f"Curation workbook -> {a.out} ({len(rows)} rows)")
        return

    # Full mode: edit a copy of the source workbook in place (source cols A-F: Name, Description,
    # Barcode, Product type, Brand name, Image URL). We fill B + F, normalise two header names to
    # Hike's exact casing, append the dimension + fenced curation columns, and colour each row by tier.
    wb = openpyxl.load_workbook(a.src)
    ws = wb.worksheets[0]
    ws.cell(row=1, column=4, value="Product Type")   # Hike field casing for a clean 1:1 import map
    ws.cell(row=1, column=5, value="Brand Name")
    for off, h in enumerate(["Depth", "Width", "Height", "Weight", DIVIDER_HEADER] + QA_HEADERS):
        ws.cell(row=1, column=7 + off, value=h)      # cols G..P
    for col in range(1, len(ALL_HEADERS) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    filled_d = filled_i = 0
    for p in rows:
        r = p["row"]
        dobj = desc_obj(desc, r)
        if dobj.get("description"):
            ws.cell(row=r, column=2, value=dobj["description"])
            filled_d += 1
        rec = img.get(r)
        url = img_url(rec)
        ws.cell(row=r, column=IMG_COL, value=url)
        if url:
            filled_i += 1
        for off, val in enumerate(dims_cells(dobj)):           # Depth/Width/Height/Weight -> G..J
            ws.cell(row=r, column=7 + off, value=val)
        for off, val in enumerate(qa_cells(rec)):              # Status..Image Flag -> L..P (K = divider)
            ws.cell(row=r, column=DIVIDER_COL + 1 + off, value=val)
        link_cell(ws.cell(row=r, column=IMG_COL), url)   # Image URL -> clickable
        set_text_source(ws, r, rec)                      # Text Source -> source domain, clickable
        set_best_guess(ws, r, rec)                       # Best guess (blanks) -> clickable
        fill = row_fill(rec)                             # whole-row tier colour
        for col in range(1, len(ALL_HEADERS) + 1):
            ws.cell(row=r, column=col).fill = fill
    fit_and_wrap(ws, len(ALL_HEADERS), ws.max_row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(ALL_HEADERS))}{ws.max_row}"
    wb.save(a.out)
    print(f"Enriched workbook -> {a.out}: {filled_d} descriptions, {filled_i} image URLs")


if __name__ == "__main__":
    main()
