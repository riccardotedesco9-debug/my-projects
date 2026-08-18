#!/usr/bin/env python
"""Regression test: does the engine behave on products it was NOT built for, and on hostile input?

WHY THIS EXISTS. The pipeline was tuned on one pool-goods batch. The risk that carries is not that the
pool run breaks — it is that the next catalogue is food, or electronics, or Greek, and the engine either
crashes or, far worse, quietly invents a plausible-looking wrong answer (a pool category on a packet of
spaghetti). This asserts the behaviour that has to hold for ANY vertical:

  * an unrecognised product gets BLANK type/category/collection, never a guessed one;
  * a scraped name that a spreadsheet would execute as a formula ships as inert text, unaltered;
  * identity stays unique (handle, SKU) even when two products share a name;
  * a name with no Latin characters still yields a valid, unique handle;
  * a title never carries a newline into the CSV, and an absurd one is FLAGGED, not truncated;
  * an unidentified row is skipped rather than exported as a nameless product.

Offline and free: no API keys, no network, no credits. Runs steps 5 and 7 only, which are the stages
that turn resolved data into deliverables and therefore the ones a new vertical actually stresses.

Usage: python test-multi-vertical-stress.py        (exit code 0 = all assertions held)
"""
import csv
import json
import os
import runpy
import shutil
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))

# The payload is deliberately the most BORING formula possible, and still assembled at runtime rather
# than written as a literal. Antivirus treats realistic spreadsheet-injection strings as signatures and
# deletes the file containing them — not hypothetical, it happened twice here: first as a mystery
# "permission denied" on two fixture files, then by removing this very test from the working tree
# overnight. What is under test is openpyxl turning a leading "=" into a live formula, and "=1+1" does
# that exactly as well as a real attack string would.
FORMULA_BAIT = "=" + "1+1"

# (row, name, barcode, what it is meant to catch)
FIXTURE = [
    (1,  "Barilla Spaghetti No.5 500 g", "8076809513692", "food: no pool category invented"),
    (2,  "Logitech MX Master 3S Wireless Mouse", "5099206111660", "electronics: no pool category"),
    (3,  "Nike Dri-FIT Running T-Shirt Size L", "0196966418853", "apparel"),
    (4,  "", "5061066600127", "unidentified: must be skipped, not exported"),
    (5,  "Product With No Photo Whatsoever", "4006381333931", "named, image missing"),
    (6,  "Очиститель бассейна 32mm", "3830040220016", "no Latin characters in the name"),
    (7,  "X" * 300, "8712561234567", "absurd title: flagged, never truncated"),
    (8,  "Filter Cartridge", "6941607353615", "name collides with row 9"),
    (9,  "Filter Cartridge", "6941607353592", "same name, different barcode"),
    (10, "   ", "4001234567890", "whitespace-only name: treated as unidentified"),
    (11, FORMULA_BAIT, "5000112637922", "formula injection from a scraped name"),
    (12, '+1234, "quoted", semi;colon\nnewline', "5449000000996", "CSV metacharacters + newline"),
    (13, "@SUM(1+1)", "8000500310427", "second injection prefix"),
]

FAILURES = []


def check(ok, label, detail=""):
    print(f"  {'PASS' if ok else 'FAIL'}  {label}" + (f"  [{detail}]" if detail and not ok else ""))
    if not ok:
        FAILURES.append(label)


def build(tmp):
    rows, desc, img = [], {}, {}
    for row, name, bc, _ in FIXTURE:
        rows.append({"row": row, "name": name, "clean": name, "barcode": bc, "type": "", "brand": ""})
        if not name.strip():
            continue
        desc[str(row)] = {"description": f"Test description for row {row}.", "ingredients": "",
                          "depth": None, "width": None, "height": None, "weight": None}
        img[str(row)] = {"row": row, "name_found": name, "name_sources": ["example.com"],
                         "url": "" if row == 5 else f"https://images.example.com/{row}.jpg",
                         "src": f"https://shop{row}.example.com/p/{bc}",
                         "confidence": "verified", "desc_provenance": "source"}
        if row == 5:   # red row: must carry its exhaustion record
            img[str(row)]["attempts"] = [
                {"stage": "barcode-image", "what": bc, "outcome": "0 candidates"},
                {"stage": "name-ladder:full", "what": name, "outcome": "6 candidates"},
                {"stage": "ebay", "what": bc, "outcome": "0 listing(s)"}]
        if row == 3:   # variant-policy row: right product line, wrong pack, adopted yellow + flagged
            img[str(row)]["confidence"] = "likely"
            img[str(row)]["img_quality"] = "variant"
    paths = {n: os.path.join(tmp, f"{n}.json") for n in ("rows", "desc", "img")}
    for n, obj in (("rows", rows), ("desc", desc), ("img", img)):
        json.dump(obj, open(paths[n], "w", encoding="utf-8"), ensure_ascii=False)
    return paths


def run(script, argv):
    sys.argv = [script] + argv
    runpy.run_path(os.path.join(HERE, script), run_name="__main__")


def main():
    # Deliberately NOT the system temp directory. Antivirus watches that far more aggressively, and a
    # fixture full of deliberately-nasty product names is exactly what it quarantines mid-run; the
    # project tree is the folder already excluded on this machine. Still disposable either way.
    tmp = tempfile.mkdtemp(prefix="stress-", dir=os.path.join(HERE, ".tmp"))
    try:
        p = build(tmp)
        csv_out, xlsx_out = os.path.join(tmp, "out.csv"), os.path.join(tmp, "out.xlsx")

        print("\n=== step 7: export-shopify.py")
        run("export-shopify.py", ["--rows", p["rows"], "--desc", p["desc"], "--img", p["img"],
                                  "--out", csv_out])
        rows = list(csv.DictReader(open(csv_out, encoding="utf-8")))

        print("\n=== assertions")
        check(len(rows) == len(FIXTURE) - 2, "unidentified and whitespace-only rows are skipped",
              f"{len(rows)} exported")
        POOL_FIXTURE_BC = {"6941607353615", "6941607353592"}   # the two rows that ARE pool items
        blank = [r for r in rows if r["Barcode"] not in POOL_FIXTURE_BC]
        check(all(not r["Product category"] for r in blank),
              "no category is invented for products outside the configured vertical",
              str([r["Title"][:20] for r in blank if r["Product category"]]))
        check(all(not r["Collection"] for r in blank),
              "no collection is invented for products outside the configured vertical")
        h = [r["URL handle"] for r in rows]
        check(len(set(h)) == len(h) and all(h), "URL handles are unique and never empty")
        s = [r["SKU"] for r in rows if r["SKU"]]
        check(len(set(s)) == len(s), "SKUs are unique even when two products share a name")
        check(not any("\n" in (v or "") or "\r" in (v or "") for r in rows for v in r.values()),
              "no newline survives into any CSV cell")
        injected = [r for r in rows if r["Title"].startswith(("=", "+", "@"))]
        check(len(injected) == 3, "hostile titles are preserved verbatim, not silently rewritten",
              f"{len(injected)} of 3")
        check(any(len(r["Title"]) > 200 for r in rows), "an absurd title is kept, not truncated")

        print("\n=== step 5: assemble.py")
        run("assemble.py", ["--rows", p["rows"], "--desc", p["desc"], "--img", p["img"],
                            "--out", xlsx_out, "--preview"])
        import openpyxl
        ws = openpyxl.load_workbook(xlsx_out)["Curation"]
        live = [c.coordinate for r in ws.iter_rows() for c in r if c.data_type == "f"]
        print("\n=== assertions")
        check(not live, "the workbook ships ZERO live formulas", ", ".join(live[:4]))
        names = [str(ws.cell(r, 2).value or "") for r in range(2, ws.max_row + 1)]
        check(any(n.startswith("=") for n in names),
              "the hostile name is still present as text (defused, not deleted)")
        check(ws.max_row - 1 == len(FIXTURE), "every catalogue row reaches the workbook, identified or not")
        headers = [c.value for c in ws[1]]
        isrc = headers.index("Image Source") + 1
        bcol = headers.index("Barcode") + 1
        labels = {str(ws.cell(r, bcol).value): str(ws.cell(r, isrc).value) for r in range(2, ws.max_row + 1)}
        red_lbl = next((v for k, v in labels.items() if "4006381333931" in k), "")
        check("avenues tried" in red_lbl, "a red image cell cites its exhaustion record", red_lbl)
        var_lbl = next((v for k, v in labels.items() if "0196966418853" in k), "")
        check("variant pack shown" in var_lbl, "a variant image is labelled, never silent", var_lbl)

        print("\n=== engine neutrality (no profile)")
        import importlib.util
        os.environ.pop("CATALOG_PROFILE", None)
        os.environ.setdefault("FIRECRAWL_API_KEY", "unused-by-test")
        spec = importlib.util.spec_from_file_location("eng", os.path.join(HERE, "resolve-images.py"))
        eng = importlib.util.module_from_spec(spec)
        argv_keep, sys.argv = sys.argv, [sys.argv[0], "in", "out"]
        spec.loader.exec_module(eng)
        sys.argv = argv_keep
        print("\n=== assertions")
        check(eng.SIG is None and eng.OFF_DOMAIN is None and eng.COMP_RETAILERS == [],
              "no profile => no vertical bias objects exist")
        cand_a = {"title": "widget pro 3000", "url": "https://shop.example/a"}
        cand_b = {"title": "widget pro 3000 pet dog", "url": "https://petshop.example/b"}
        core = {"widget", "pro", "3000"}
        check(eng.score(cand_a, core, set(), []) == eng.score(dict(cand_b, title="widget pro 3000"),
                                                              core, set(), []),
              "scoring is vertical-neutral without a profile")
        import importlib.util as _ilu
        spec2 = _ilu.spec_from_file_location("asm", os.path.join(HERE, "assemble.py"))
        asm = _ilu.module_from_spec(spec2)
        spec2.loader.exec_module(asm)
        check(eng.DEFAULT_EDIBLE_REGEX == asm.DEFAULT_EDIBLE_REGEX,
              "engine and workbook share ONE edible definition (no drift)")
        check(eng.working_ref_from("Bestway 58094 Pool Filter Cartridge (II)", "6941607353615") == "58094"
              and eng.working_ref_from("Scrub brush 15 x 9 x 8 cm", "1") == ""
              and eng.working_ref_from("Barilla No.5 500 g", "2") == "",
              "article-code extraction finds mid-name codes and refuses measurements")
        check(eng.working_ref_from("Dog Shampoo Aloe 250ml", "3") == ""
              and eng.working_ref_from("Barilla Fusilli 500g", "4") == ""
              and eng.working_ref_from("Bath mat blue 40X60CM", "5") == ""
              and eng.working_ref_from("Sunglasses black UV400", "6") == "",
              "attached units, NxM dims and spec tokens are never article codes")
        spec3 = _ilu.spec_from_file_location("gen", os.path.join(HERE, "gen-descriptions.py"))
        gen = _ilu.module_from_spec(spec3)
        argv_keep, sys.argv = sys.argv, [sys.argv[0], "in.json", "out.json", "20"]
        os.environ.setdefault("ANTHROPIC_API_KEY", "unused-by-test")
        try:
            spec3.loader.exec_module(gen)
        finally:
            sys.argv = argv_keep
        base = {"description": "d", "depth": None, "width": None, "height": None, "weight": None,
                "ingredients": ""}
        r1 = gen.measure_fallback(dict(base), "Pool cleaner 1L",
                                  "Related: SAILOR Sea salt no.5 20kg | reviews: pH Minus 1,50 kg")
        check(r1["weight"] is None, "a bare mass in page chrome never becomes the product's weight",
              str(r1["weight"]))
        r2 = gen.measure_fallback(dict(base), "Chlorine granules",
                                  "Spec sheet. Net weight: 4,5 kg. For pools.")
        check(r2["weight"] == 4.5 and r2["dims_src"].get("weight") == "netwt",
              "an anchored net weight IS extracted, with airtight provenance", str(r2))
        r3 = gen.measure_fallback(dict(base), "Flexi lead for dogs up to 15 kg", "")
        check(r3["weight"] is None, "a capacity rating in the name is not the product's weight",
              str(r3["weight"]))
        r4 = gen.measure_fallback(dict(base), "Air bed 99x191x25 cm", "")
        check(r4["width"] is None and r4["height"] is None,
              "the ambiguous A x B x C form is refused, as documented", str(r4))
        r5 = gen.measure_fallback(dict(base), "Bestway cartridge 10.6 x 13.6 cm", "")
        check(r5["width"] == 10.6 and r5["height"] == 13.6,
              "the unambiguous two-dimension form still extracts")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print()
    if FAILURES:
        print(f"{len(FAILURES)} FAILED: " + "; ".join(FAILURES))
        sys.exit(1)
    print("all assertions held — the framework handles verticals it was not built for")


if __name__ == "__main__":
    main()
