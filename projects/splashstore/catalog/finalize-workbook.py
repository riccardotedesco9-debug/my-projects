#!/usr/bin/env python
"""Finalize the curation workbook for publishing: surface everything the run learned, in the Sheet.

WHY. The engine's curation workbook is a review surface for the fields the ENGINE resolves (identity,
photo, description, ingredients, dimensions). Three kinds of relevant information sit outside that
and were previously invisible to whoever opens the Sheet:

  1. Shopify's own fields (Tags, SEO) — real data, but they lived only in a CSV on disk, so a reviewer
     reading the Sheet reasonably concluded they had not been produced at all.
  2. The provenance behind a GREEN row — how many independent domains confirmed the barcode, and the
     page it was confirmed on. "Verified" is more trustworthy when you can see who verified it.
  3. The scanned codes that are NOT product barcodes. They were dropped before the engine ran, so the
     workbook showed 44 rows for 51 scanned items. Silently shrinking the list makes the job look more
     finished than it is — the same reason `read-scans.py` refuses to write a lossy catalogue.

So this adds columns to Curation, a Shopify tab, and a tab for the codes that need re-scanning.
Idempotent: re-running rebuilds rather than appending duplicates.

Usage: python finalize-workbook.py [workbook.xlsx] [products.csv] [skipped.json] [resolved.json] [catalog.json]
"""
import csv
import json
import os
import sys

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

WORKBOOK = sys.argv[1] if len(sys.argv) > 1 else "../docs/splashstore-scan-curation-260816.xlsx"
CSV_PATH = sys.argv[2] if len(sys.argv) > 2 else "../docs/splashstore-shopify-import-260816.csv"
SKIPPED = sys.argv[3] if len(sys.argv) > 3 else "../.tmp/catalog-skipped.json"
RESOLVED = sys.argv[4] if len(sys.argv) > 4 else "../.tmp/resolved.json"
CATALOG = sys.argv[5] if len(sys.argv) > 5 else "../.tmp/catalog.json"
DESC = sys.argv[6] if len(sys.argv) > 6 else "../.tmp/desc.json"

HEAD = PatternFill("solid", fgColor="00975A")        # matches the engine's own header green
TODO = PatternFill("solid", fgColor="FFF4CC")        # yours to fill, not an engine miss
GREY = PatternFill("solid", fgColor="EDEDED")
HEAD_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="0563C1", underline="single")

# Tinted = the human still has to supply it. SKU and Inventory are no longer here: SKU is generated
# and Inventory carries the owner's 999 placeholder, so tinting them would flag filled data as a gap.
TODO_COLS = {"Price", "Compare-at price", "Cost per item"}
WIDE = {"Title": 44, "Description": 60, "Tags": 38, "SEO title": 34, "SEO description": 50,
        "URL handle": 34, "Product image URL": 30, "Image alt text": 30}

KIND_NOTE = {
    "gs1-serial": "GS1 serial capture (AI 21/250) — identifies one physical unit, carries no product code",
    "in-store":   "in-store / variable-weight prefix — not globally unique, cannot be verified",
    "other":      "supplier or model code, not a GTIN",
}

# Only what the engine's own columns do NOT already say. Deliberately small:
#   SEO title / SEO description duplicate Name / Description (identical in 41/41 rows) and Shopify
#     falls back to those anyway — they live on the Shopify tab for the owner to fill, not here.
#   "Confirmed on page" duplicated Description Source, whose cell is ALREADY a hyperlink to that URL.
#   URL handle is derived from Name and belongs with the import schema, not the review view.
# What survives is genuinely new: the tag set, and HOW MANY independent domains confirmed the barcode
# (the tier label says "Verified", not whether one source or three agreed).
EXTRA = ["Tags", "Confirmed by"]

# Review order: triage signal first (photo + status), then identity and the evidence for it, then the
# enriched content, then measurements, then raw source URLs. Ingredients sit late — for a non-edible
# vertical they are N/A on almost every row and would otherwise split the useful columns in half.
CURATION_ORDER = [
    "Image", "Status", "Name", "Barcode", "Confirmed by",
    "Brand Name", "Product Type",
    "Description", "Description Source", "Tags",
    "Depth (cm)", "Width (cm)", "Height (cm)", "Weight (kg)",
    "Image URL", "Image Source",
    "Ingredients", "Ingredients Source",
    "Best guess (manual)",
]


# Column widths by HEADER NAME, applied after the reorder. They have to be re-asserted by name
# because openpyxl's delete_cols moves cells but NOT column_dimensions, so every re-run of
# enrich_curation shifted the widths one place out of step with their headers: Description ended up
# 14 wide while its Source column had 56, which is what made the text stack into tall thin ribbons.
# Sized for reading: the two long prose fields get real room, the numeric ones stay narrow.
CURATION_WIDTHS = {
    "Status": 10, "Name": 42, "Barcode": 16, "Confirmed by": 26,
    "Brand Name": 18, "Product Type": 20,
    "Description": 62, "Description Source": 30, "Tags": 30,
    "Depth (cm)": 10, "Width (cm)": 10, "Height (cm)": 10, "Weight (kg)": 10,
    "Image URL": 26, "Image Source": 30,
    "Ingredients": 34, "Ingredients Source": 24,
    "Best guess (manual)": 22,
}


def apply_widths(ws, widths):
    """Set column widths by header name, so they cannot drift out of step with their columns."""
    for i, h in enumerate([c.value for c in ws[1]], start=1):
        if h in widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[h]


def reorder_columns(ws, order):
    """Rearrange columns to `order`, carrying values, styling, hyperlinks and widths.

    Column A is never moved: the embedded product thumbnails are anchored to A2..An, and anchors do
    not follow a column rewrite — reordering A would leave every photo pointing at the wrong row.
    """
    headers = [c.value for c in ws[1]]
    target = [h for h in order if h in headers] + [h for h in headers if h not in order]
    if target[0] != headers[0]:
        raise SystemExit(f"refusing to move column A ({headers[0]}) — image anchors would break")
    if target == headers:
        return False

    def snapshot(col):
        return [{
            "value": ws.cell(r, col).value,
            "fill": ws.cell(r, col).fill.copy(),
            "font": ws.cell(r, col).font.copy(),
            "align": ws.cell(r, col).alignment.copy(),
            "fmt": ws.cell(r, col).number_format,
            "link": ws.cell(r, col).hyperlink.target if ws.cell(r, col).hyperlink else None,
        } for r in range(1, ws.max_row + 1)]

    cols = {h: snapshot(i + 1) for i, h in enumerate(headers)}
    widths = {h: ws.column_dimensions[get_column_letter(i + 1)].width
              for i, h in enumerate(headers)}

    for new_i, h in enumerate(target, start=1):
        for r, cell_data in enumerate(cols[h], start=1):
            # `ws.cell(r, c, value)` SKIPS the assignment when value is None (openpyxl treats the
            # third argument as "optional"), which during an in-place reorder leaves the previous
            # occupant of that position sitting there. Every column that should have gone blank
            # instead kept the old column's content — Depth showed an image URL. Assign explicitly.
            cell = ws.cell(r, new_i)
            cell.value = cell_data["value"]
            cell.fill, cell.font = cell_data["fill"], cell_data["font"]
            cell.alignment, cell.number_format = cell_data["align"], cell_data["fmt"]
            cell.hyperlink = cell_data["link"] or None
        ws.column_dimensions[get_column_letter(new_i)].width = widths[h]

    # Prove the move rather than trust it: every cell must now hold exactly what its column held
    # before. Cheap (rows x columns) and it is the only thing standing between a silent column
    # scramble and a published catalogue that looks fine until someone reads a row.
    for new_i, h in enumerate(target, start=1):
        for r, cell_data in enumerate(cols[h], start=1):
            if ws.cell(r, new_i).value != cell_data["value"]:
                raise SystemExit(f"reorder lost data: {h} row {r} became "
                                 f"{ws.cell(r, new_i).value!r}, expected {cell_data['value']!r}")
    return True


PX_PER_PT = 96 / 72          # Excel row heights are points; images are pixels
CHAR_PX = 7.0                # approx px per character at the default font
LINE_PT = 14.0               # one wrapped line of text
IMG_PAD_PX = 26              # slack around a thumbnail: Drive's xlsx->Sheets conversion does not
                             # reproduce row heights exactly, and a flush fit clips in the browser
MAX_ROW_PT = 320             # sanity ceiling so one long description cannot create a giant row


def fit_layout(ws, wrap_cols, img_col_width_px=None):
    """Size rows and columns so NOTHING is clipped — not a thumbnail, not a wrapped description.

    Two independent constraints per row, and the row must satisfy both:
      * the tallest embedded image in it, plus padding;
      * the tallest wrapped text among `wrap_cols`, estimated from the column's own width.
    Sized here rather than left to the viewer because Sheets will not auto-fit a row that contains
    an over-the-grid image, and a fixed height set for the image alone silently truncates long text.
    """
    headers = [c.value for c in ws[1]]
    widths = {h: (ws.column_dimensions[get_column_letter(i + 1)].width or 10)
              for i, h in enumerate(headers)}

    img_h = {}
    for im in getattr(ws, "_images", []):
        r = im.anchor._from.row + 1
        img_h[r] = max(img_h.get(r, 0), im.height)
    if img_h and img_col_width_px:
        need = max(im.width for im in ws._images) + IMG_PAD_PX
        ws.column_dimensions["A"].width = max(widths.get(headers[0], 10), need / CHAR_PX)

    for h in wrap_cols:
        if h not in headers:
            continue
        col = headers.index(h) + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in range(2, ws.max_row + 1):
        need_pt = ((img_h[r] + IMG_PAD_PX) / PX_PER_PT) if r in img_h else 0
        for h in wrap_cols:
            if h not in headers:
                continue
            text = str(ws.cell(r, headers.index(h) + 1).value or "")
            if not text:
                continue
            per_line = max(8, int(widths[h] * CHAR_PX / CHAR_PX))   # chars that fit on one line
            lines = max(1, -(-len(text) // per_line))
            need_pt = max(need_pt, lines * LINE_PT + 6)
        if need_pt:
            ws.row_dimensions[r].height = min(round(need_pt, 1), MAX_ROW_PT)


def style_header(ws, row=1):
    for c in ws[row]:
        if c.value is not None:
            c.fill, c.font = HEAD, HEAD_FONT
            c.alignment = Alignment(vertical="center", wrap_text=True)


def enrich_curation(ws, by_barcode, resolved_by_barcode):
    """Append Tags / SEO / provenance beside the engine's own columns, matched on Barcode."""
    headers = [c.value for c in ws[1]]
    # Rebuild on re-run. Delete BY NAME, highest index first: after reorder_columns the EXTRA columns
    # are no longer adjacent (Confirmed by sits at E, Tags at J), so deleting N consecutive columns
    # from the first one's index removed innocent neighbours — it ate "Depth (cm)".
    for name in sorted((h for h in EXTRA if h in headers),
                       key=lambda n: headers.index(n), reverse=True):
        ws.delete_cols(headers.index(name) + 1)
        headers = [c.value for c in ws[1]]
    start = len(headers) + 1
    bc_col = headers.index("Barcode") + 1

    for i, name in enumerate(EXTRA):
        ws.cell(1, start + i, name)
    for r in range(2, ws.max_row + 1):
        bc = str(ws.cell(r, bc_col).value or "").strip()
        row = by_barcode.get(bc, {})
        rec = resolved_by_barcode.get(bc, {})
        srcs = rec.get("name_sources") or []
        values = [
            row.get("Tags", ""),
            f"{len(srcs)} source{'s' if len(srcs) != 1 else ''}: " + ", ".join(
                s.replace("www.", "") for s in srcs) if srcs else "",
        ]
        for i, v in enumerate(values):
            cell = ws.cell(r, start + i, v)
            cell.alignment = Alignment(vertical="top", wrap_text=(i == 0))
        # A scanned row has no brand/type of its own; both are derived in the Shopify export from
        # words literally present in the confirmed name. Mirror them here so the review view is not
        # showing two blank columns for data that exists one tab over.
        for col, key in (("Brand Name", "Vendor"), ("Product Type", "Type")):
            if col in headers and row.get(key) and not ws.cell(r, headers.index(col) + 1).value:
                ws.cell(r, headers.index(col) + 1, row[key])
    widths = {"Tags": 34, "Confirmed by": 30}
    for i, name in enumerate(EXTRA):
        ws.column_dimensions[get_column_letter(start + i)].width = widths[name]
    style_header(ws)


def build_shopify_tab(wb, rows):
    if "Shopify import" in wb.sheetnames:
        del wb["Shopify import"]
    ws = wb.create_sheet("Shopify import")
    headers = rows[0]
    for r_i, row in enumerate(rows, start=1):
        for c_i, value in enumerate(row, start=1):
            cell = ws.cell(r_i, c_i, value)
            if r_i > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=headers[c_i - 1] in WIDE)
                if headers[c_i - 1] in TODO_COLS:
                    cell.fill = TODO
    # Barcodes must stay text: a spreadsheet renders 700175992406 as 7.00176E+11 and a round-trip
    # would corrupt the one globally-unique field we have.
    if "Barcode" in headers:
        col = get_column_letter(headers.index("Barcode") + 1)
        for r_i in range(2, len(rows) + 1):
            ws[f"{col}{r_i}"].number_format = "@"
    for c_i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c_i)].width = WIDE.get(h, max(12, min(len(h) + 4, 22)))
    ws.freeze_panes = "C2"
    style_header(ws)
    return len(rows) - 1


def build_unscannable_tab(wb, skipped):
    """The scanned codes that carry no product barcode. They are real stock the owner still has to
    deal with; leaving them out of the deliverable would quietly understate the remaining work."""
    if "Needs re-scan" in wb.sheetnames:
        del wb["Needs re-scan"]
    ws = wb.create_sheet("Needs re-scan")
    for i, h in enumerate(["Scanned code", "What it is", "What to do"], start=1):
        ws.cell(1, i, h)
    for r, s in enumerate(skipped, start=2):
        ws.cell(r, 1, s["scanned"]).number_format = "@"
        ws.cell(r, 2, KIND_NOTE.get(s["kind"], s["kind"])).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 3, "Re-scan the product's own EAN/UPC from the packaging, or enter it by hand.")
        for c in range(1, 4):
            ws.cell(r, c).fill = GREY
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 52
    ws.freeze_panes = "A2"
    style_header(ws)
    return len(skipped)


def check_thumbnails(ws, resolved, catalog):
    """Every row that HAS an image URL must carry an embedded thumbnail.

    A row can end up with a URL but no picture — the image was re-resolved and its normalized JPG was
    left stale or deleted, or the fallback fetch failed at assemble time. The cell then renders as a
    broken-image icon in the Sheet, which looks like a bug in the data rather than a missing file.
    Loud here, because it is invisible once published.
    """
    headers = [c.value for c in ws[1]]
    bc_col = headers.index("Barcode") + 1
    have = {im.anchor._from.row + 1 for im in getattr(ws, "_images", [])}
    by_barcode = {c["barcode"]: str(c["row"]) for c in catalog if c.get("barcode")}

    missing = []
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, bc_col).value or "").strip()
        key = by_barcode.get(code)
        if key and (resolved.get(key, {}) or {}).get("url") and r not in have:
            missing.append((r, code))
    # And the other direction, which is the dangerous one: a photo shown on a row the catalogue says
    # has NO image. A missing thumbnail looks broken and gets noticed; a surplus one looks like data
    # and gets believed, even when it is a candidate the engine already rejected.
    orphan = []
    for r in range(2, ws.max_row + 1):
        key = by_barcode.get(str(ws.cell(r, bc_col).value or "").strip())
        if r in have and not (resolved.get(key, {}) or {}).get("url"):
            orphan.append(r)
    if missing:
        print("  !! rows with an image URL but NO embedded thumbnail (will render broken):")
        for r, code in missing:
            print(f"       sheet row {r}  barcode {code}  -> delete .tmp/normalized/<row>.jpg and re-run step 4")
    if orphan:
        print(f"  !! {len(orphan)} row(s) DISPLAY a photo but hold no image URL "
              f"(sheet rows {', '.join(map(str, orphan))}) -> a stale .tmp/normalized/<row>.jpg, "
              f"or a best-guess being embedded. The row would show a picture it does not own.")
    return len(missing) + len(orphan)


# The colour contract, stated once so a run can be checked against it rather than eyeballed:
#   GREEN  = verified by one of the implemented factual methods (the GTIN or a brand article code
#            found LITERALLY at the source). Never a name match, never a vision guess.
#   YELLOW = we have something, but it was not confirmed by one of those methods. Review it.
#   RED    = we have nothing for this field.
#   GREY   = the field does not apply (ingredients on a non-edible).
# Colours are per FIELD, not per row: an image can be green while the description beside it is yellow.
# The row's Status is just the weakest field present.
COLOUR_BY_RGB = {"C6EFCE": "green", "FFEB9C": "yellow", "FFC7CE": "red", "808080": "grey"}
BARCODE_TIERS = {"verified", "verified-cross", "verified-official"}


def cell_colour(cell):
    """Normalised colour name. The alpha prefix differs between a locally written workbook (00…) and
    one round-tripped through Drive (FF…), so compare on the last six hex digits only — an earlier
    version of this check compared full strings, silently matched nothing, and reported a clean bill
    of health for a workbook it had never actually inspected."""
    rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
    return COLOUR_BY_RGB.get(str(rgb)[-6:].upper(), "")


def check_colour_rule(ws, resolved, desc, catalog):
    """Assert the colour contract on every enriched field. Returns a list of violations.

    Worth enforcing in code because the colour is the ONLY thing a reviewer reads before trusting a
    row: a green cell that is actually a guess is worse than no colour at all.
    """
    headers = [c.value for c in ws[1]]
    by_barcode = {c["barcode"]: str(c["row"]) for c in catalog if c.get("barcode")}
    bc_col = headers.index("Barcode") + 1
    violations = []

    for r in range(2, ws.max_row + 1):
        key = by_barcode.get(str(ws.cell(r, bc_col).value or "").strip())
        rec = (resolved.get(key) or {}) if key else {}
        d = (desc.get(key) or {}) if key else {}
        conf = str(rec.get("confidence") or "")
        for field, src_col, has_value, verified in (
            ("image", "Image Source", bool(rec.get("url")), conf in BARCODE_TIERS),
            ("description", "Description Source", bool(d.get("description")),
             rec.get("desc_provenance") == "source"),
            ("ingredients", "Ingredients Source", bool(d.get("ingredients")),
             str(rec.get("ingredients_src", "")).startswith(("source", "verified", "cross", "official"))),
        ):
            if src_col not in headers:
                continue
            colour = cell_colour(ws.cell(r, headers.index(src_col) + 1))
            if colour == "green" and not (has_value and verified):
                violations.append((r, field, "GREEN without a factual confirmation"))
            elif colour == "red" and has_value:
                violations.append((r, field, "RED but a value exists"))
            elif colour == "yellow" and not has_value:
                violations.append((r, field, "YELLOW but empty"))
    return violations


# Hover notes on the header cells. The colour contract is obvious once you know it and baffling until
# you do — a green image beside a yellow description looks like a bug rather than the design. Drive
# converts xlsx comments into Sheets notes, so these travel with the published sheet.
HEADER_NOTES = {
    "Status": ("Worst field in the row, not a judgement on the whole product.\n"
               "READY = every field present is green.\n"
               "REVIEW = at least one field is yellow.\n"
               "HOLD = at least one field is missing (red).\n"
               "A row can be REVIEW while its photo is fully verified."),
    "Image Source": ("Colour applies to THIS FIELD ONLY.\n"
                     "GREEN = the barcode was found literally on the page the photo came from.\n"
                     "YELLOW = the right product, but the photo needs a look (crop, multipack, unclear).\n"
                     "RED = no usable photo was found.\n"
                     "The cell names the site the photo came from. Compare it with Description Source: "
                     "a different site there means the two fields were proven separately."),
    "Description Source": ("Colour applies to THIS FIELD ONLY, and is judged separately from the photo.\n"
                           "GREEN = written from the barcode-confirmed page for this exact product.\n"
                           "YELLOW = written from a name-matched page, so it may describe a variant.\n"
                           "RED = no description.\n"
                           "A yellow description beside a green image is normal: the photo was proven "
                           "by the barcode, the wording was not.\n"
                           "The cell names the page the wording came from, or says 'generic (from name)' "
                           "when there was no page at all."),
    "Ingredients Source": ("Colour applies to THIS FIELD ONLY.\n"
                           "GREY = not applicable (the product is not edible).\n"
                           "GREEN = composition from a barcode-confirmed source.\n"
                           "YELLOW = composition from a brand or retailer page that was not "
                           "barcode-confirmed.\nRED = missing on an edible product."),
    "Confirmed by": ("How many INDEPENDENT domains printed this barcode.\n"
                     "Three sources agreeing is stronger evidence than one, though both are green."),
}

# The same reminder as one line, parked in the header row just past the last column. It lives ON the
# table on purpose: the question it answers ("why is the photo green and the description yellow?") is
# asked while reading a row, and a separate explainer tab is somewhere you have to remember to go. An
# earlier version was a whole "How to read this" tab; that was more sheet than the point deserved.
NOTE_TEXT = ("Colours are per field, not per row. Green: the barcode was found literally at that "
             "source. Yellow: plausible, but not confirmed that way. Red: nothing. Grey: not "
             "applicable. A green photo beside a yellow description is normal. "
             "Hover any header for detail.")
NOTE_MARKER = "Colours are per field"
LEGACY_TAB = "How to read this"


def strip_note(ws):
    """Drop a note left by a previous run, so the header row parses as pure column names again.

    Everything downstream reads `ws[1]` as the header list, so a leftover sentence there would be
    handled as if it were a column. Deleting the whole column is safe precisely because the note is
    always the last one: nothing sits to its right, so no cell, hyperlink or image anchor moves.
    """
    for c in ws[1]:
        if isinstance(c.value, str) and c.value.startswith(NOTE_MARKER):
            ws.delete_cols(c.column)
            return True
    return False


def explain_red_images(ws, resolved, catalog):
    """Hover the engine's actual verdict onto a red Image Source cell.

    A red cell says a photo is missing. It cannot say whether the web had nothing to offer, or whether
    a candidate WAS found and the image AI judged it a different product. That difference decides what
    the owner does next — search harder, or go photograph the shelf — and the engine already recorded
    it, so put it where the question gets asked.
    """
    headers = [c.value for c in ws[1]]
    if "Image Source" not in headers:
        return 0
    col = headers.index("Image Source") + 1
    bc_col = headers.index("Barcode") + 1
    by_barcode = {c["barcode"]: str(c["row"]) for c in catalog if c.get("barcode")}
    added = 0
    for r in range(2, ws.max_row + 1):
        rec = resolved.get(by_barcode.get(str(ws.cell(r, bc_col).value or "").strip(), ""), {}) or {}
        if rec.get("url"):
            continue
        # EVERY red row explains itself, not only the ones vision rejected. A blank cell with no note
        # is the worst of all outcomes: it reads as an engine that gave up silently, when these are
        # four different situations with four different next actions.
        if rec.get("vision_verdict"):
            note = ("A photo WAS found and checked against this product's confirmed name.\n"
                    "The image AI judged it a different product, so it was not used:\n\n"
                    + str(rec["vision_verdict"])[:400]
                    + "\n\nThis one wants photographing in the garage, not more searching.")
        elif rec.get("best_url"):
            note = ("A candidate photo was found but never judged, so it was not used.\n\n"
                    "Best candidate: " + str(rec.get("best_title") or rec["best_url"])[:130]
                    + "\n\nIt is linked under 'Best guess (manual)' if you want to look at it.")
        elif not rec.get("name_found"):
            note = ("No source anywhere confirmed this barcode, so the product has no name — and with "
                    "no name there is nothing to search an image with.\n\nIdentify it by hand: check "
                    "the packaging, or search the code on the brand's own site.")
        else:
            note = ("The product IS identified from its barcode, but no usable photo was found by any "
                    "source we search: the barcode databases, the barcode directories, an image search "
                    "on the confirmed name, and eBay.\n\nNothing left to try automatically. Photograph "
                    "it in the garage.")
        tried = rec.get("attempts") or []
        if tried:
            listing = "\n".join(f"- {a.get('stage')}: {a.get('what')} -> {a.get('outcome')}"
                                 for a in tried[:10])
            note += f"\n\nAvenues tried ({len(tried)}):\n" + listing
        ws.cell(r, col).comment = Comment(note, "catalogue engine", height=240, width=420)
        added += 1
    return added


TIER_FILL = {"verified": PatternFill("solid", fgColor="C6EFCE"),
             "likely": PatternFill("solid", fgColor="FFEB9C"),
             "blank": PatternFill("solid", fgColor="FFC7CE"),
             "na": PatternFill("solid", fgColor="EDEDED")}


def colour_derived(ws, resolved, desc, catalog):
    """Colour every DERIVED cell by its evidence, so no unverified value hides behind white.

    The engine's own three fields (Image / Description / Ingredients) were always coloured; Name,
    Brand, Product Type, Tags and the four measurements rendered white regardless of evidence, so a
    value that was merely derived or model-extracted looked exactly like a confirmed one. Mapping:
      Name          green = barcode-confirmed; red = a GTIN row nothing could identify;
                    white = the catalogue supplied it (a POS export input, not an enrichment).
      Brand/Type/   follow the name they were derived from (green name -> green, else yellow);
      Tags          red when an identified row derived nothing; grey when unidentified (nothing to
                    derive FROM — the red already lives on Name).
      Dims/Weight   green when the number was literally stated (page text or the confirmed name),
                    yellow when only the model produced it, grey when not stated anywhere.
    Row Status stays worst-of Image/Description/Ingredients — this is visual truth, not a new
    status input, so today's READY/REVIEW/HOLD counts cannot shift under the owner.
    """
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers) if h}
    by_barcode = {c["barcode"]: c for c in catalog if c.get("barcode")}
    painted = 0
    for r in range(2, ws.max_row + 1):
        row = by_barcode.get(str(ws.cell(r, col["Barcode"]).value or "").strip())
        if row is None:
            continue
        key = str(row["row"])
        rec = resolved.get(key, {}) or {}
        d = desc.get(key, {}) or {}
        scan = not (row.get("name") or "").strip() or bool(rec.get("name_found"))
        identified = bool(rec.get("name_found") or (row.get("name") or "").strip())
        if scan:
            name_tier = "verified" if rec.get("name_found") else "blank"
        else:
            name_tier = None      # catalogue input, not an enrichment — stays white
        if name_tier and "Name" in col:
            ws.cell(r, col["Name"]).fill = TIER_FILL[name_tier]
            painted += 1
        for h in ("Brand Name", "Product Type", "Tags"):
            if h not in col:
                continue
            has = bool(str(ws.cell(r, col[h]).value or "").strip())
            if not scan and has:
                continue          # POS-supplied input
            t = ("verified" if name_tier in (None, "verified") else "likely") if has \
                else ("blank" if identified else "na")
            ws.cell(r, col[h]).fill = TIER_FILL[t]
            painted += 1
        dsrc = d.get("dims_src") or {}
        for h, f in (("Depth (cm)", "depth"), ("Width (cm)", "width"),
                     ("Height (cm)", "height"), ("Weight (kg)", "weight")):
            if h not in col:
                continue
            has = ws.cell(r, col[h]).value not in (None, "")
            # GREEN only for text whose provenance is airtight: the engine-written Net-weight prefix
            # (GTIN-keyed) or the confirmed NAME. A "page" match is anchored but the page itself may
            # be a likely-tier page, so it reviews as yellow like any other unverified specific.
            t = ("verified" if dsrc.get(f) in ("netwt", "name") else "likely") if has else "na"
            ws.cell(r, col[h]).fill = TIER_FILL[t]
            painted += 1
    return painted


def annotate_variant_images(ws, resolved, catalog):
    """Hover note on a yellow image adopted under the variant policy: the photo shows the right
    product line in a DIFFERENT pack size, by explicit owner decision — usable, flagged, never green."""
    headers = [c.value for c in ws[1]]
    if "Image Source" not in headers:
        return 0
    col = headers.index("Image Source") + 1
    bc = headers.index("Barcode") + 1
    by_barcode = {c["barcode"]: str(c["row"]) for c in catalog if c.get("barcode")}
    added = 0
    for r in range(2, ws.max_row + 1):
        rec = resolved.get(by_barcode.get(str(ws.cell(r, bc).value or "").strip(), ""), {}) or {}
        if rec.get("url") and rec.get("img_quality") == "variant":
            ws.cell(r, col).comment = Comment(
                "The image AI judged this photo to be the SAME product line in a DIFFERENT pack "
                "size or count than this barcode's. Adopted as yellow by policy: a recognisable "
                "variant photo beats an empty cell, but verify the size before it goes live.",
                "catalogue engine", height=150, width=360)
            added += 1
    return added


def defuse_formulas(wb):
    """Re-assert the engine's no-live-formula rule across every sheet, and report the count.

    `assemble.py` already defuses the workbook it writes, but this script REWRITES cells: the column
    reorder re-assigns each value, and the Shopify tab is built fresh from the CSV. Both paths hand
    openpyxl a raw string, so a scraped name beginning with "=" becomes a formula again on the way out.
    This runs last, immediately before the save that gets published.
    """
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str):
                    cell.data_type = "s"
                    n += 1
    return n


def add_note(ws):
    """Write the colour reminder one column past the table, in the frozen header row.

    Narrow column and no wrap, so the text spills across the empty cells beside it and reads as a
    margin note rather than a 19th column of data.
    """
    cell = ws.cell(1, ws.max_column + 1, NOTE_TEXT)
    cell.font = Font(italic=True, color="7F7F7F", size=9)
    cell.alignment = Alignment(vertical="center")
    ws.column_dimensions[get_column_letter(cell.column)].width = 3
    return cell.coordinate


def annotate_headers(ws):
    """Attach the per-field notes to the header cells, where the confusion actually happens."""
    headers = [c.value for c in ws[1]]
    added = 0
    for name, note in HEADER_NOTES.items():
        if name in headers:
            cell = ws.cell(1, headers.index(name) + 1)
            cell.comment = Comment(note, "catalogue engine", height=170, width=340)
            added += 1
    return added


def main():
    wb = openpyxl.load_workbook(WORKBOOK)
    with open(CSV_PATH, encoding="utf-8") as f:
        rows = list(csv.reader(f))
    headers = rows[0]
    by_barcode = {r[headers.index("Barcode")]: dict(zip(headers, r)) for r in rows[1:]}

    catalog = json.load(open(CATALOG, encoding="utf-8"))
    resolved = json.load(open(RESOLVED, encoding="utf-8"))
    resolved_by_barcode = {c["barcode"]: resolved.get(str(c["row"]), {}) for c in catalog}
    skipped = json.load(open(SKIPPED, encoding="utf-8"))

    # Before anything reads row 1 as headers.
    strip_note(wb["Curation"])
    if LEGACY_TAB in wb.sheetnames:
        del wb[LEGACY_TAB]

    enrich_curation(wb["Curation"], by_barcode, resolved_by_barcode)
    moved = reorder_columns(wb["Curation"], CURATION_ORDER)
    n_shop = build_shopify_tab(wb, rows)
    n_skip = build_unscannable_tab(wb, skipped)
    broken = check_thumbnails(wb["Curation"], resolved, catalog)
    desc = json.load(open(DESC, encoding="utf-8")) if os.path.exists(DESC) else {}
    violations = check_colour_rule(wb["Curation"], resolved, desc, catalog)
    for r, field, why in violations:
        print(f"  !! colour rule: sheet row {r} {field} — {why}")

    # Widths first: fit_layout estimates how many lines a cell wraps to from its column width, so a
    # width fixed afterwards would leave every row height computed against the wrong number.
    apply_widths(wb["Curation"], CURATION_WIDTHS)

    # Last: size everything so no photo and no sentence is clipped in the browser.
    # The two Source columns belong in here as well: they wrap like any other text, and leaving them
    # out meant their rows were sized from the OTHER columns only, so a two-line provenance label sat
    # in a row built for one line.
    fit_layout(wb["Curation"],
               ["Name", "Description", "Tags", "Confirmed by", "Ingredients", "Best guess (manual)",
                "Image Source", "Description Source", "Ingredients Source"],
               img_col_width_px=True)
    fit_layout(wb["Shopify import"],
               ["Title", "Description", "Tags", "Collection", "Image alt text", "URL handle"])
    fit_layout(wb["Needs re-scan"], ["What it is", "What to do"])
    notes = annotate_headers(wb["Curation"])
    where = add_note(wb["Curation"])
    explained = explain_red_images(wb["Curation"], resolved, catalog)
    painted = colour_derived(wb["Curation"], resolved, desc, catalog)
    variants = annotate_variant_images(wb["Curation"], resolved, catalog)
    defused = defuse_formulas(wb)
    wb.save(WORKBOOK)

    print(f"Finalized {WORKBOOK}")
    print(f"  Curation      + {', '.join(EXTRA)}"
          f"{' | reordered for review flow' if moved else ''}")
    print(f"  Shopify import  {n_shop} products x {len(headers)} columns "
          f"({', '.join(sorted(TODO_COLS & set(headers)))} tinted for you)")
    print(f"  Needs re-scan   {n_skip} codes that carry no product barcode")
    print(f"  thumbnails      {'ALL PRESENT' if not broken else str(broken) + ' MISSING - see above'}")
    print(f"  colour guide    inline note at {where} + {notes} header hover notes (no explainer tab)")
    print(f"  colour rule     {'HOLDS (green = factually verified only)' if not violations else str(len(violations)) + ' VIOLATIONS - see above'}")
    print(f"  red images      {explained} carry the verdict that rejected their candidate")
    print(f"  derived colour  {painted} cell(s) now show their evidence tier (no white unknowns)"
          + (f" | {variants} variant image(s) flagged" if variants else ""))
    print(f"  formula safety  {'no live formulas' if not defused else str(defused) + ' scraped cell(s) forced to text'}")
    print(f"  total scanned accounted for: {wb['Curation'].max_row - 1} + {n_skip} = "
          f"{wb['Curation'].max_row - 1 + n_skip}")


if __name__ == "__main__":
    main()
